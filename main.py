import pandas as pd
import openpyxl
from typing import Union, Any
from numpy import ndarray
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame
from pandas.core.indexing import _LocIndexer
from sklearn import preprocessing, metrics
from sklearn.neighbors import KNeighborsClassifier
from sklearn.model_selection import train_test_split


def read_file() -> Union[dict[Any, DataFrame], DataFrame]:
    df = pd.read_excel('DataSet_TB3.xlsx')
    return df


def select_column_name(df: Union[dict[Any, DataFrame], DataFrame]):
    columns_name = df.loc[:, ~df.columns.isin(['label'])].columns
    return columns_name


def get_rows_values(df: Union[dict[Any, DataFrame], DataFrame], columns_name: _LocIndexer) -> ndarray:
    rows_data = df[columns_name].values
    return rows_data


def get_rows_label_values(df: Union[dict[Any, DataFrame], DataFrame]) -> ndarray:
    rows_label_data = df['label'].values
    return rows_label_data


def split_dataset(rows_data: ndarray, label_data: ndarray, test_size: float):
    x_train, x_test, y_train, y_test = train_test_split(rows_data, label_data, test_size=test_size)
    return x_train, x_test, y_train, y_test


def get_rows_data_without_id(data):
    return data[:, 1:]


def get_rows_only_id(data):
    return data[:, 0]


def create_classifier(x_data, y_data, k: int, metric='minkowski') -> KNeighborsClassifier:
    knn = KNeighborsClassifier(n_neighbors=k, metric=metric).fit(x_data, y_data)
    return knn


def predict_data(knn: KNeighborsClassifier, prediction_data):
    return knn.predict(prediction_data)


def calculate_accuracy(actual_result, test_result) -> float:
    return metrics.accuracy_score(actual_result, test_result)


def grouping_result(test_result, test_result_id, actual_result) -> list[list[int, int]]:
    result = []

    for i in range(0, len(test_result_id)):
        idData = test_result_id[i]
        prediction = test_result[i]
        actual = actual_result[i]
        result.append([idData, prediction, actual])

    return result


def open_workbook(k: int, file_name='OutputValidasi.xlsx') -> Workbook:
    try:
        wb = openpyxl.load_workbook(filename=file_name)
    except:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'k={k}'
    return wb


def get_sheet(workbook: Workbook, name: str) -> Worksheet:
    try:
        sheet = workbook[name]
    except:
        sheet = workbook.create_sheet(name)
    return sheet


def write_result_group(result_group: list, k: int, test_accuracy: float):
    workbook = open_workbook(k)
    sheet = get_sheet(workbook, f'k={k}')

    sheet['A1'] = 'idData'
    sheet['B1'] = 'Label Aktual'
    sheet['C1'] = 'Hasil Klasifikasi'
    sheet['D1'] = 'Akurasi'

    sheet['D2'] = test_accuracy*100

    for i in range(0, len(result_group)):
        rowAddress = 2 + i
        result = result_group[i]
        idData = result[0]
        prediction = result[1]
        actual = result[2]

        sheet[f'A{rowAddress}'] = idData
        sheet[f'B{rowAddress}'] = actual
        sheet[f'C{rowAddress}'] = prediction

    workbook.save('OutputValidasi.xlsx')
    workbook.close()


def do_knn(k: int, metric_method='minkowski', test_size=0.1):
    print('-'*40)
    print(f'Creating report for k={k} with method {metric_method}')
    print('-'*40)

    print(f'Reading dataset from file...')
    df = read_file()
    columns_name = select_column_name(df)
    rows_data = get_rows_values(df, columns_name)
    rows_label_data = get_rows_label_values(df)
    print(f'Reading dataset from file complete')

    print(f'Splitting dataset with {test_size} of dataset as test...')
    x_train, x_test, y_train, y_test = split_dataset(rows_data, rows_label_data, test_size)
    print(f'Splitting dataset with {test_size} of dataset as test complete')

    print(f'Separating the id and attribute from dataset...')
    x_train_no_id = get_rows_data_without_id(x_train)
    x_test_no_id = get_rows_data_without_id(x_test)
    x_test_id = get_rows_only_id(x_test)
    print(f'Separating the id and attribute from dataset complete.')

    print('Creating classifier using train data..')
    knn = create_classifier(x_train_no_id, y_train, k, metric_method)
    print('Creating classifier using train data complete.')

    print('Generating prediction result for test data...')
    test_predict_result = predict_data(knn, x_test_no_id)
    print('Generating prediction result for test data complete.')

    print('Calculating test accuracy...')
    test_accuracy = calculate_accuracy(y_test, test_predict_result)
    print('Calculating test accuracy complete.')

    print('Grouping test result...')
    test_result_group = grouping_result(test_predict_result, x_test_id, y_test)
    print('Grouping test result complete.')

    print('Writing to Output File...')
    write_result_group(test_result_group, k, test_accuracy)
    print('Writing to Output File complete.')

    print('-' * 40)
    print(f'End of report k={k} with method {metric_method}')
    print('-' * 40)
    print(' ')


if __name__ == '__main__':
    do_knn(3, 'manhattan')
    do_knn(5)
    do_knn(7)
    do_knn(9, 'euclidean')
    do_knn(2)
